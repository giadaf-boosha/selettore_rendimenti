"""
Comparison Exporter - Export Excel per report di confronto.

Genera file Excel formattati per i confronti fondi vs ETF
con formattazione condizionale per i delta e foglio riepilogo.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import Optional
from io import BytesIO
from datetime import datetime
import logging

from core.models import ComparisonReport
from exporters.styles import (
    create_header_style,
    create_thin_border,
    get_performance_font,
    get_alternate_row_fill,
)

logger = logging.getLogger(__name__)


class ComparisonExporter:
    """
    Esporta report confronto in Excel formattato.

    Genera file con:
    - Foglio "Confronto": tabella principale con performance e delta
    - Foglio "Riepilogo": statistiche e metriche aggregate
    - Formattazione condizionale: verde per outperformance, rosso per underperformance
    """

    # Colonne tabella confronto
    BASE_COLUMNS = [
        ("Nome", 45),
        ("ISIN", 14),
        ("Tipo", 8),
        ("Origine", 10),
        ("Categoria", 30),
    ]

    PERFORMANCE_COLUMNS = [
        ("Perf. 1m", 10),
        ("Perf. 3m", 10),
        ("Perf. 6m", 10),
        ("Perf. YTD", 10),
        ("Perf. 1a", 10),
        ("Perf. 3a", 10),
        ("Perf. 5a", 10),
        ("Perf. 7a", 10),
        ("Perf. 9a", 10),
        ("Perf. 10a", 10),
    ]

    DELTA_COLUMNS = [
        ("Delta 1m", 10),
        ("Delta 3m", 10),
        ("Delta 6m", 10),
        ("Delta YTD", 10),
        ("Delta 1a", 10),
        ("Delta 3a", 10),
        ("Delta 5a", 10),
        ("Delta 7a", 10),
        ("Delta 9a", 10),
        ("Delta 10a", 10),
    ]

    def __init__(self):
        """Inizializza l'exporter."""
        self.header_style = create_header_style()
        self.border = create_thin_border()
        self.alt_row_fill = get_alternate_row_fill()

        # Fill condizionali per delta
        self.green_fill = PatternFill(
            start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'
        )
        self.red_fill = PatternFill(
            start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
        )

    def export(self, report: ComparisonReport, filename: Optional[str] = None) -> BytesIO:
        """
        Esporta report confronto in Excel.

        Args:
            report: ComparisonReport da esportare
            filename: Nome file (opzionale, per metadata)

        Returns:
            BytesIO buffer con file Excel
        """
        logger.info(f"Exporting comparison report: {len(report.results)} results")

        wb = Workbook()

        # Foglio Confronto
        ws_comparison = wb.active
        ws_comparison.title = "Confronto"
        self._create_comparison_sheet(ws_comparison, report)

        # Foglio Riepilogo
        ws_summary = wb.create_sheet("Riepilogo")
        self._create_summary_sheet(ws_summary, report, filename)

        # Foglio Benchmark (se disponibile)
        if report.benchmark_etf:
            ws_benchmark = wb.create_sheet("Benchmark")
            self._create_benchmark_sheet(ws_benchmark, report)

        # Salva in buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info("Comparison export completed")
        return buffer

    def _create_comparison_sheet(self, ws, report: ComparisonReport) -> None:
        """Crea foglio principale con tabella confronto."""
        # Determina colonne da includere
        all_columns = self.BASE_COLUMNS + self.PERFORMANCE_COLUMNS + self.DELTA_COLUMNS

        # Scrivi header
        for col_idx, (header_name, _) in enumerate(all_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.header_style['border']

        # Scrivi dati
        for row_idx, result in enumerate(report.results, start=2):
            inst = result.instrument

            # Colonne base
            ws.cell(row=row_idx, column=1, value=inst.name)
            ws.cell(row=row_idx, column=2, value=inst.isin)
            ws.cell(row=row_idx, column=3, value=inst.instrument_type.value)
            ws.cell(row=row_idx, column=4, value=result.origin.capitalize())
            ws.cell(row=row_idx, column=5, value=inst.category_morningstar or inst.category_assogestioni or "")

            # Colonne performance
            perf_values = [
                inst.perf_1m_eur,
                inst.perf_3m_eur,
                inst.perf_6m_eur,
                inst.perf_ytd_eur,
                inst.perf_1y_eur,
                inst.perf_3y_eur,
                inst.perf_5y_eur,
                inst.perf_7y_eur,
                inst.perf_9y_eur,
                inst.perf_10y_eur,
            ]

            for col_offset, perf_value in enumerate(perf_values):
                col_idx = 6 + col_offset
                cell = ws.cell(row=row_idx, column=col_idx)
                if perf_value is not None:
                    cell.value = perf_value / 100
                    cell.number_format = '0.00%'
                    cell.font = get_performance_font(perf_value >= 0)
                else:
                    cell.value = ""

            # Colonne delta (solo per strumenti universo)
            delta_values = [
                result.delta_1m,
                result.delta_3m,
                result.delta_6m,
                result.delta_ytd,
                result.delta_1y,
                result.delta_3y,
                result.delta_5y,
                result.delta_7y,
                result.delta_9y,
                result.delta_10y,
            ]

            for col_offset, delta_value in enumerate(delta_values):
                col_idx = 16 + col_offset  # Dopo le colonne performance
                cell = ws.cell(row=row_idx, column=col_idx)
                if delta_value is not None and result.origin == "universe":
                    cell.value = delta_value / 100
                    cell.number_format = '+0.00%;-0.00%'
                    # Colore condizionale
                    if delta_value > 0.5:
                        cell.fill = self.green_fill
                    elif delta_value < -0.5:
                        cell.fill = self.red_fill
                else:
                    cell.value = "" if result.origin == "market" else ""

            # Bordi e formattazione riga
            for col_idx in range(1, len(all_columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.border

                # Righe alternate
                if row_idx % 2 == 0 and not cell.fill.fgColor or cell.fill.fgColor.rgb == '00000000':
                    cell.fill = self.alt_row_fill

        # Imposta larghezze colonne
        for col_idx, (_, width) in enumerate(all_columns, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # Filtri automatici
        if report.results:
            last_col = get_column_letter(len(all_columns))
            ws.auto_filter.ref = f"A1:{last_col}{len(report.results) + 1}"

        # Blocca prima riga
        ws.freeze_panes = "A2"

    def _create_summary_sheet(
        self,
        ws,
        report: ComparisonReport,
        filename: Optional[str] = None
    ) -> None:
        """Crea foglio riepilogo con statistiche."""
        # Titolo
        ws.cell(row=1, column=1, value="Riepilogo Confronto").font = Font(bold=True, size=14)

        # Metadata
        metadata = [
            ("Data confronto", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Tipo confronto", report.comparison_type.replace("_", " ").title()),
            ("Categoria", report.category or "N/A"),
            ("Sistema categorizzazione", (report.category_type or "N/A").title()),
            ("", ""),  # Riga vuota
            ("Strumenti totali", report.total_instruments),
            ("Strumenti universo", report.universe_count),
            ("ETF mercato", report.market_count),
            ("", ""),
            ("Outperformer (3a)", report.outperformers_count),
            ("Underperformer (3a)", report.underperformers_count),
            ("", ""),
        ]

        # Aggiungi medie delta
        for period, avg in report.avg_delta.items():
            label = f"Media Delta {period}"
            value = f"{avg:+.2f}%"
            metadata.append((label, value))

        # Best/Worst performer
        metadata.append(("", ""))
        if report.best_performer:
            best_name = report.best_performer.instrument.name[:40]
            best_delta = report.best_performer.delta_3y
            metadata.append(("Best Performer", best_name))
            metadata.append(("  Delta 3a", f"{best_delta:+.2f}%" if best_delta else "N/A"))

        if report.worst_performer:
            worst_name = report.worst_performer.instrument.name[:40]
            worst_delta = report.worst_performer.delta_3y
            metadata.append(("Worst Performer", worst_name))
            metadata.append(("  Delta 3a", f"{worst_delta:+.2f}%" if worst_delta else "N/A"))

        # Scrivi metadata
        for row_idx, (key, value) in enumerate(metadata, start=3):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True) if key else Font()
            ws.cell(row=row_idx, column=2, value=value)

        # Imposta larghezze
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _create_benchmark_sheet(self, ws, report: ComparisonReport) -> None:
        """Crea foglio con dettagli ETF benchmark."""
        etf = report.benchmark_etf
        if not etf:
            return

        # Titolo
        ws.cell(row=1, column=1, value="ETF Benchmark").font = Font(bold=True, size=14)

        # Dettagli
        details = [
            ("Nome", etf.name),
            ("ISIN", etf.isin),
            ("Tipo", etf.instrument_type.value),
            ("Valuta", etf.currency),
            ("Domicilio", etf.domicile or "N/A"),
            ("Distribuzione", etf.distribution.value),
            ("Categoria Morningstar", etf.category_morningstar or "N/A"),
            ("Categoria Assogestioni", etf.category_assogestioni or "N/A"),
            ("", ""),
            ("Performance", ""),
            ("  1 mese", f"{etf.perf_1m_eur:.2f}%" if etf.perf_1m_eur else "N/A"),
            ("  3 mesi", f"{etf.perf_3m_eur:.2f}%" if etf.perf_3m_eur else "N/A"),
            ("  6 mesi", f"{etf.perf_6m_eur:.2f}%" if etf.perf_6m_eur else "N/A"),
            ("  YTD", f"{etf.perf_ytd_eur:.2f}%" if etf.perf_ytd_eur else "N/A"),
            ("  1 anno", f"{etf.perf_1y_eur:.2f}%" if etf.perf_1y_eur else "N/A"),
            ("  3 anni", f"{etf.perf_3y_eur:.2f}%" if etf.perf_3y_eur else "N/A"),
            ("  5 anni", f"{etf.perf_5y_eur:.2f}%" if etf.perf_5y_eur else "N/A"),
            ("  7 anni", f"{etf.perf_7y_eur:.2f}%" if etf.perf_7y_eur else "N/A"),
            ("  9 anni", f"{etf.perf_9y_eur:.2f}%" if etf.perf_9y_eur else "N/A"),
            ("  10 anni", f"{etf.perf_10y_eur:.2f}%" if etf.perf_10y_eur else "N/A"),
            ("", ""),
            ("Fonti dati", ", ".join(etf.sources) if etf.sources else "N/A"),
            ("Qualita' dati", f"{etf.data_quality_score:.0f}%"),
        ]

        for row_idx, (key, value) in enumerate(details, start=3):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True) if key and not key.startswith(" ") else Font()
            ws.cell(row=row_idx, column=2, value=value)

        # Imposta larghezze
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def export_to_file(self, report: ComparisonReport, filepath: str) -> str:
        """
        Esporta direttamente su file.

        Args:
            report: ComparisonReport da esportare
            filepath: Percorso file di destinazione

        Returns:
            Percorso file salvato
        """
        buffer = self.export(report)

        with open(filepath, 'wb') as f:
            f.write(buffer.getvalue())

        logger.info(f"Comparison Excel saved to {filepath}")
        return filepath


def comparison_to_dataframe(report: ComparisonReport) -> pd.DataFrame:
    """
    Converte ComparisonReport in DataFrame.

    Args:
        report: ComparisonReport da convertire

    Returns:
        DataFrame pandas
    """
    if not report or not report.results:
        return pd.DataFrame()

    data = []
    for result in report.results:
        inst = result.instrument
        row = {
            "Nome": inst.name,
            "ISIN": inst.isin,
            "Tipo": inst.instrument_type.value,
            "Origine": result.origin.capitalize(),
            "Categoria": inst.category_morningstar or inst.category_assogestioni or "",
            "Perf. 1m": inst.perf_1m_eur,
            "Perf. 3m": inst.perf_3m_eur,
            "Perf. 6m": inst.perf_6m_eur,
            "Perf. YTD": inst.perf_ytd_eur,
            "Perf. 1a": inst.perf_1y_eur,
            "Perf. 3a": inst.perf_3y_eur,
            "Perf. 5a": inst.perf_5y_eur,
            "Perf. 7a": inst.perf_7y_eur,
            "Perf. 9a": inst.perf_9y_eur,
            "Perf. 10a": inst.perf_10y_eur,
        }

        # Aggiungi delta solo per strumenti universo
        if result.origin == "universe":
            row.update({
                "Delta 1m": result.delta_1m,
                "Delta 3m": result.delta_3m,
                "Delta 6m": result.delta_6m,
                "Delta YTD": result.delta_ytd,
                "Delta 1a": result.delta_1y,
                "Delta 3a": result.delta_3y,
                "Delta 5a": result.delta_5y,
                "Delta 7a": result.delta_7y,
                "Delta 9a": result.delta_9y,
                "Delta 10a": result.delta_10y,
            })

        data.append(row)

    return pd.DataFrame(data)
