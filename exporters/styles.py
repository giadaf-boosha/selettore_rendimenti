"""
Stili Excel per formattazione professionale.

Definisce colori, font, bordi e altri stili riutilizzabili
per la generazione di file Excel.
"""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle

# Palette colori corporate
COLORS = {
    'primary_dark': '1F4E79',      # Blu scuro header
    'primary_light': '5B9BD5',     # Blu chiaro
    'white': 'FFFFFF',
    'black': '000000',
    'gray_light': 'F2F2F2',        # Righe alternate
    'gray_medium': 'D9D9D9',
    'green_dark': '006400',        # Performance positiva
    'green_light': 'C6EFCE',       # Background performance positiva
    'red_dark': '8B0000',          # Performance negativa
    'red_light': 'FFC7CE',         # Background performance negativa
    'yellow_warning': 'FFD966',    # Warning
}


def create_header_style() -> dict:
    """
    Crea stile per header tabella.

    Returns:
        Dict con componenti stile
    """
    return {
        'font': Font(name='Calibri', size=11, bold=True, color=COLORS['white']),
        'fill': PatternFill(start_color=COLORS['primary_dark'],
                           end_color=COLORS['primary_dark'], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': create_thin_border(),
    }


def create_thin_border() -> Border:
    """
    Crea bordo sottile standard.

    Returns:
        Border object
    """
    thin_side = Side(border_style='thin', color=COLORS['black'])
    return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def create_data_style() -> dict:
    """
    Crea stile per celle dati.

    Returns:
        Dict con componenti stile
    """
    return {
        'font': Font(name='Calibri', size=10),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': create_thin_border(),
    }


def create_number_style() -> dict:
    """
    Crea stile per celle numeriche.

    Returns:
        Dict con componenti stile
    """
    return {
        'font': Font(name='Calibri', size=10),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': create_thin_border(),
    }


def get_performance_font(positive: bool) -> Font:
    """
    Restituisce font per performance positiva/negativa.

    Args:
        positive: True se performance positiva

    Returns:
        Font colorato appropriatamente
    """
    color = COLORS['green_dark'] if positive else COLORS['red_dark']
    return Font(name='Calibri', size=10, color=color)


def get_performance_fill(positive: bool) -> PatternFill:
    """
    Restituisce fill per performance positiva/negativa.

    Args:
        positive: True se performance positiva

    Returns:
        PatternFill colorato appropriatamente
    """
    color = COLORS['green_light'] if positive else COLORS['red_light']
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def get_alternate_row_fill() -> PatternFill:
    """
    Restituisce fill per righe alternate.

    Returns:
        PatternFill grigio chiaro
    """
    return PatternFill(
        start_color=COLORS['gray_light'],
        end_color=COLORS['gray_light'],
        fill_type='solid'
    )


# Larghezze colonne consigliate (v3.0 con periodi estesi)
COLUMN_WIDTHS = {
    'Nome': 45,
    'ISIN': 14,
    'Tipo': 8,
    'Valuta': 8,
    'Domicilio': 12,
    'Distribuzione': 12,
    'Cat. Morningstar': 30,
    'Cat. Assogestioni': 30,
    'Perf. 1m': 10,
    'Perf. 3m': 10,
    'Perf. 6m': 10,
    'Perf. YTD': 10,
    'Perf. 1a': 10,
    'Perf. 3a': 10,
    'Perf. 5a': 10,
    'Perf. 7a': 10,
    'Perf. 9a': 10,
    'Perf. 10a': 10,
    'Delta 1m': 10,
    'Delta 3m': 10,
    'Delta 6m': 10,
    'Delta YTD': 10,
    'Delta 1a': 10,
    'Delta 3a': 10,
    'Delta 5a': 10,
    'Delta 7a': 10,
    'Delta 9a': 10,
    'Delta 10a': 10,
    'Volatilita\' 3a': 12,
    'Sharpe 3a': 10,
    'Fonti': 20,
    'Qualita\' Dati': 12,
}
