"""
Excel Writer - Generazione file Excel formattati professionalmente.

Crea file Excel con:
- Header colorati
- Formattazione condizionale per performance
- Filtri automatici
- Righe alternate
- Foglio metadata
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Optional, Union
from io import BytesIO
from datetime import datetime
import logging

from core.models import AggregatedInstrument
from exporters.styles import (
    COLORS,
    COLUMN_WIDTHS,
    create_header_style,
    create_thin_border,
    get_performance_font,
    get_alternate_row_fill,
)

logger = logging.getLogger(__name__)


class ExcelWriter:
    """
    Genera file Excel formattati professionalmente.

    Output conforme ai requisiti RF-05 delle specifiche:
    - Nome, ISIN, Categorie, Valuta
    - Performance YTD, 1a, 3a, 5a, 7a, 10a
    - Fonte dati
    """

    # Colonne da includere nel file (v3.0 con periodi estesi)
    COLUMNS = [
        ("Nome", 45),
        ("ISIN", 14),
        ("Tipo", 8),
        ("Valuta", 8),
        ("Distribuzione", 12),
        ("Cat. Morningstar", 30),
        ("Cat. Assogestioni", 30),
        ("Perf. 1m", 10),
        ("Perf. 3m", 10),
        ("Perf. 6m", 10),
        ("Perf. YTD", 10),
        ("Perf. 1a", 10),
        ("Perf. 3a", 10),
        ("Perf. 5a", 10),
        ("Perf. 7a", 10),
        ("Perf. 9a", 10),
        ("Perf. 10a", 10),
        ("Fonti", 20),
    ]

    # Indici colonne performance (0-based, relativo a COLUMNS)
    PERF_COLUMNS = [7, 8, 9, 10, 11, 12, 13, 14, 15, 16]

    def __init__(self):
        """Inizializza l'Excel writer."""
        self.header_style = create_header_style()
        self.border = create_thin_border()
        self.alt_row_fill = get_alternate_row_fill()

    def export(
        self,
        instruments: List[AggregatedInstrument],
        filename: Optional[str] = None
    ) -> BytesIO:
        """
        Esporta lista strumenti in Excel formattato.

        Args:
            instruments: Lista di strumenti aggregati
            filename: Nome file (opzionale, per header)

        Returns:
            BytesIO buffer con file Excel
        """
        logger.info(f"Exporting {len(instruments)} instruments to Excel")

        # Crea workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Risultati"

        # Scrivi header
        self._write_headers(ws)

        # Scrivi dati
        self._write_data(ws, instruments)

        # Applica formattazione
        self._apply_formatting(ws, len(instruments))

        # Auto-dimensiona colonne
        self._auto_fit_columns(ws)

        # Aggiungi filtri automatici
        self._add_auto_filter(ws, len(instruments))

        # Crea foglio metadata
        self._create_metadata_sheet(wb, len(instruments), filename)

        # Salva in buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info("Excel export completed")
        return buffer

    def _write_headers(self, ws) -> None:
        """Scrive e formatta la riga header."""
        for col_idx, (header_name, _) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.header_style['border']

    def _write_data(self, ws, instruments: List[AggregatedInstrument]) -> None:
        """Scrive i dati nelle celle."""
        for row_idx, inst in enumerate(instruments, start=2):
            # Nome
            ws.cell(row=row_idx, column=1, value=inst.name)
            # ISIN
            ws.cell(row=row_idx, column=2, value=inst.isin)
            # Tipo
            ws.cell(row=row_idx, column=3, value=inst.instrument_type.value)
            # Valuta
            ws.cell(row=row_idx, column=4, value=inst.currency)
            # Distribuzione
            ws.cell(row=row_idx, column=5, value=inst.distribution.value)
            # Categoria Morningstar
            ws.cell(row=row_idx, column=6, value=inst.category_morningstar or "")
            # Categoria Assogestioni
            ws.cell(row=row_idx, column=7, value=inst.category_assogestioni or "")

            # Performance con formato percentuale (v3.0 con periodi estesi)
            perf_values = [
                inst.perf_1m_eur,
                inst.perf_3m_eur,
                inst.perf_6m_eur,
                inst.perf_ytd_eur,
                inst.perf_1y_eur,
                inst.perf_3y_eur,
                inst.perf_5y_eur,
                inst.perf_7y_eur,
                inst.perf_9y_eur,
                inst.perf_10y_eur,
            ]

            for col_offset, perf_value in enumerate(perf_values):
                cell = ws.cell(row=row_idx, column=8 + col_offset)
                if perf_value is not None:
                    # Converti in decimale per formato percentuale
                    cell.value = perf_value / 100
                    cell.number_format = '0.00%'
                    # Colore condizionale
                    cell.font = get_performance_font(perf_value >= 0)
                else:
                    cell.value = ""

            # Fonti
            ws.cell(row=row_idx, column=18, value=", ".join(inst.sources))

    def _apply_formatting(self, ws, row_count: int) -> None:
        """Applica formattazione alle righe dati."""
        for row_idx in range(2, row_count + 2):
            # Righe alternate
            if row_idx % 2 == 0:
                for col_idx in range(1, len(self.COLUMNS) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Non sovrascrivere fill delle performance
                    if col_idx not in [c + 1 for c in self.PERF_COLUMNS]:
                        cell.fill = self.alt_row_fill

            # Bordi a tutte le celle
            for col_idx in range(1, len(self.COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).border = self.border

            # Allineamento
            for col_idx in range(1, len(self.COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx in [c + 1 for c in self.PERF_COLUMNS]:
                    cell.alignment = Alignment(horizontal='right')
                elif col_idx == 1:  # Nome
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='center')

    def _auto_fit_columns(self, ws) -> None:
        """Imposta larghezza colonne."""
        for col_idx, (col_name, width) in enumerate(self.COLUMNS, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width

    def _add_auto_filter(self, ws, row_count: int) -> None:
        """Aggiunge filtri automatici all'header."""
        if row_count > 0:
            last_col_letter = ws.cell(row=1, column=len(self.COLUMNS)).column_letter
            ws.auto_filter.ref = f"A1:{last_col_letter}{row_count + 1}"

    def _create_metadata_sheet(
        self,
        wb: Workbook,
        result_count: int,
        filename: Optional[str] = None
    ) -> None:
        """Crea foglio con metadati dell'export."""
        ws_meta = wb.create_sheet("Info")

        metadata = [
            ("Generato il", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Strumenti totali", result_count),
            ("Applicazione", "Selettore Rendimenti Fondi/ETF"),
            ("Versione", "1.0.0"),
            ("Valuta Performance", "EUR"),
        ]

        if filename:
            metadata.append(("Nome file", filename))

        for row_idx, (key, value) in enumerate(metadata, start=1):
            ws_meta.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            ws_meta.cell(row=row_idx, column=2, value=value)

        ws_meta.column_dimensions['A'].width = 25
        ws_meta.column_dimensions['B'].width = 35

    def export_to_file(
        self,
        instruments: List[AggregatedInstrument],
        filepath: str
    ) -> str:
        """
        Esporta direttamente su file.

        Args:
            instruments: Lista strumenti
            filepath: Percorso file di destinazione

        Returns:
            Percorso file salvato
        """
        buffer = self.export(instruments)

        with open(filepath, 'wb') as f:
            f.write(buffer.getvalue())

        logger.info(f"Excel saved to {filepath}")
        return filepath


def instruments_to_dataframe(instruments: List[AggregatedInstrument]) -> pd.DataFrame:
    """
    Converte lista strumenti in DataFrame.

    Args:
        instruments: Lista di AggregatedInstrument

    Returns:
        DataFrame pandas
    """
    data = []
    for inst in instruments:
        data.append({
            "Nome": inst.name,
            "ISIN": inst.isin,
            "Tipo": inst.instrument_type.value,
            "Valuta": inst.currency,
            "Distribuzione": inst.distribution.value,
            "Cat. Morningstar": inst.category_morningstar or "",
            "Cat. Assogestioni": inst.category_assogestioni or "",
            "Perf. 1m": inst.perf_1m_eur,
            "Perf. 3m": inst.perf_3m_eur,
            "Perf. 6m": inst.perf_6m_eur,
            "Perf. YTD": inst.perf_ytd_eur,
            "Perf. 1a": inst.perf_1y_eur,
            "Perf. 3a": inst.perf_3y_eur,
            "Perf. 5a": inst.perf_5y_eur,
            "Perf. 7a": inst.perf_7y_eur,
            "Perf. 9a": inst.perf_9y_eur,
            "Perf. 10a": inst.perf_10y_eur,
            "Volatilita' 3a": inst.volatility_3y,
            "Sharpe 3a": inst.sharpe_ratio_3y,
            "Fonti": ", ".join(inst.sources),
            "Qualita'": f"{inst.data_quality_score:.0f}%",
        })

    return pd.DataFrame(data)
